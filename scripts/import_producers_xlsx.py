"""One-shot import: load mehamekor_producers_final.xlsx into the dev DB.

Uses the same `import_rows` function that powers POST /admin/producers/import,
so the column mapping (CLAUDE.md / admin_brief.docx §4) is identical.
"""
import os
import sys

os.environ.setdefault(
    "DATABASE_URL", "postgresql://postgres:postgres@localhost:5432/mehamakor"
)
sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", "backend"))

from openpyxl import load_workbook

from app.database import SessionLocal
from app.services.producer_import import import_rows

XLSX = os.path.join(os.path.dirname(__file__), "..", "mehamekor_producers_final.xlsx")

wb = load_workbook(XLSX, data_only=True)
ws = wb.active
rows = list(ws.iter_rows(min_row=2, values_only=True))
print(f"Loaded {len(rows)} data rows from {ws.title}")

db = SessionLocal()
try:
    summary = import_rows(db, rows, dry_run=False)
finally:
    db.close()

print(f"\nImported: {summary['imported']}")
print(f"Skipped:  {summary['skipped']}")
print(f"Errors:   {summary['errors']}")
print("\n--- per-row ---")
for r in summary["rows"]:
    flag = "OK " if not r["errors"] else "ERR"
    warns = f"  ⚠ {'; '.join(r['warnings'])}" if r["warnings"] else ""
    errs = f"  ✗ {'; '.join(r['errors'])}" if r["errors"] else ""
    print(f"  row {r['row_number']:>3} {flag}  {r['data'].get('name')!r:<40} {r['data'].get('city') or '':<15}{warns}{errs}")
