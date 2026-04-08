"""Enrich mehamekor_producers_updated.xlsx with data scraped from producer
websites (gathered via web search). Only fills EMPTY cells — never overwrites.

Output: mehamekor_producers_final.xlsx
"""
from pathlib import Path
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent.parent
SRC = ROOT / "mehamekor_producers_updated.xlsx"
DST = ROOT / "mehamekor_producers_final.xlsx"

# Column letters → indices (1-based, openpyxl)
COL = {c: i + 1 for i, c in enumerate("ABCDEFGHIJKLMNOPQRSTUVW")}

# Per-row enrichment payload, keyed by spreadsheet row number.
# Each value is a dict {column_letter: value}. Only applied where cell is empty.
ENRICHMENTS = {
    7: {  # משק חביבאן — havivian.co.il
        "D": "https://www.instagram.com/meshek_havivian_/",
    },
    8: {  # שורשי ציון — col E only has plain text, no URL
        "H": "בית שמש",
        "I": "מוצרים מוכנים",
        "J": "כן",
        "K": "כן",
        "L": "ירושלים, מרכז",
        "N": ("שורשי ציון מייצרים מזון איכותי טבעוני, רואו-פוד, אורגני וללא "
              "גלוטן. חומרי הגלם נרכשים ישירות מחקלאים אורגניים מקומיים, "
              "ללא חימום או בישול, בכמויות קטנות ובעבודת יד — קרקרים, "
              "שוקולד, עוגיות, גרנולה, קינוחים ואגוזים מונבטים."),
        "O": "מזון רואו אורגני וטבעוני בעבודת יד",
        "Q": "shorshei-tzion",
        "U": "לא",
        "V": "כן",
        "W": "אתר: en.shoresheitzion.com — בית שמש, רחוב הצבע 3",
    },
    9: {  # משק הר הפרחים — ecomeshek.co.il
        "C": "054-6306374",
    },
    10: {  # The Seven Fat Cows — thesevenfatcows.com
        # contact, phone not public. instagram already filled.
    },
    11: {  # סנדרה פירות איכותיים — mypips.app/sandraperot
        "B": "סנדרה",
        "C": "052-2518862",
        "D": "https://www.instagram.com/sandra_perot_/",
        "H": "פרדס חנה",
        "I": "פירות",
        "J": "כן",
        "K": "כן",
        "L": "פרדס חנה והסביבה",
        "N": ("חנות פירות וירקות איכותיים בפרדס חנה. תוצרת טרייה ישירות "
              "מהחקלאים, מהקטיף של הבוקר עד הבית. הזמנות דרך mypips, "
              "פייסבוק ואינסטגרם."),
        "O": "פירות וירקות טריים מהשדה לבית",
        "Q": "sandra-perot",
        "U": "לא",
        "V": "לא",
    },
    12: {  # הלחם של גל — mypips.app/gal.lehem
        "H": "קריית טבעון",
        "I": "לחמים ואפייה",
        "J": "כן",
        "K": "לא",
        "L": "קריית טבעון והסביבה",
        "N": ("לחמי מחמצת בעבודת יד מקמח חיטה מלא 100%. הבצק מכיל רק "
              "קמח, מים וקצת מלח ים. אפייה ביתית-מקצועית בימי שלישי, "
              "רביעי ושישי. ההזמנות מראש דרך החנות באינטרנט, איסוף "
              "עצמי מ-10 נקודות איסוף."),
        "O": "לחמי מחמצת מקמח מלא בעבודת יד",
        "Q": "gal-lehem",
        "U": "לא",
        "V": "לא",
    },
    13: {  # חוות קפדנה — kapandafarm.co.il
        "D": "https://www.instagram.com/kapandafarm/",
    },
    14: {  # חוות וקסמן — mypips.app/vaxmanfarm
        "D": "https://www.instagram.com/vaxmanfarm/",
    },
    15: {  # משק שניידר / מושבוצ — moshbutz.co.il
        "B": "משפחת שניידר",
        "D": "https://www.instagram.com/moshbutz_schneider/",
        "H": "קצרין, רמת הגולן",
        "I": "בשר ודגים",
        "J": "כן",
        "K": "כן",
        "L": "כל הארץ",
        "M": "כשר",
        "N": ("הקצביה של מושבוצ ומשק שניידר — קצביה איכותית מהגולן. "
              "הבקר גדל זמן ארוך במרעה פתוח ברמת הגולן לפי פרוטוקול ייחודי "
              "של משק שניידר, ללא זרזי גדילה וללא אנטיביוטיקה מונעת, תוך "
              "שמירה על רווחת הבקר. משלוח חינם מעל ₪500."),
        "O": "בשר בקר טרי מהגולן בגידול חופשי",
        "Q": "moshbutz-schneider",
        "U": "כן",
        "V": "לא",
    },
    16: {  # פופ — mypips.app/popisrael
        "H": "כל הארץ",
        "I": "ירקות",
        "J": "כן",
        "K": "כן",
        "L": "מרכז, צפון",
        "N": ("פופ — חוויית קנייה ישראלית-קהילתית שמשנה את הצריכה בארץ. "
              "אחרי 7 באוקטובר נולדה תנועת ערבות הדדית שמחברת בין צרכנים "
              "לחקלאים ויצרנים ישראלים שנפגעו: ירקות ופירות טריים, "
              "לחמים בעבודת יד ומוצרי איכות ישירות מהיצרן ללא מתווכים."),
        "O": "ירקות ופירות טריים ישירות מחקלאים ישראלים",
        "Q": "pop-israel",
        "U": "לא",
        "V": "לא",
    },
    17: {  # חוות קיפוד — havatkipod.co.il / mypips.app/kipodfarm
        "D": "https://www.instagram.com/kipod.farm/",
        "H": "בוויז (הוד השרון)",
        "I": "ירקות",
        "J": "כן",
        "K": "כן",
        "L": "הוד השרון, כפר סבא, רעננה, השרון הדרומי, הרצליה, רמת השרון, צפון תל אביב",
        "N": ("חוות קיפוד היא מרכז אקולוגי לאדם ולאדמה. גידול ירקות ועלים "
              "ירוקים בעבודת יד, ללא ריסוסים ודשנים כימיים. האדמה מעובדת "
              "ידנית ומועשרת בקומפוסט איכותי. דוכן בחווה בימי רביעי "
              "17:00-20:00, חמישי 11:00-20:00 ושישי 10:00-14:00."),
        "O": "ירקות ועלים ירוקים ללא ריסוסים, בעבודת יד",
        "Q": "havat-kipod",
        "U": "לא",
        "V": "לא",
        "W": "ללא ריסוסים אך לא בהסמכה אורגנית רשמית",
    },
    18: {  # סל מהטבע — דפנה אמון — mypips.app/vigenbari
        "B": "דפנה אמון",
        "H": "יפו, תל אביב",
        "I": "פירות",
        "J": "כן",
        "K": "כן",
        "L": "תל אביב והסביבה",
        "N": ("סל מהטבע של דפנה אמון ואריק אלון — שירות סלי פירות וירקות "
              "טריים מחקלאים אורגניים בכל הארץ, יחד עם מוצרים משלימים "
              "מהצומח. דפנה היא טבעונית ותיקה, חיה תזונה רואו עשירה בפירות, "
              "ומפעילה גם בלוג וערוץ יוטיוב VeganBari."),
        "O": "סלי פירות וירקות אורגניים מהטבע",
        "Q": "sal-mehateva",
        "U": "לא",
        "V": "כן",
    },
    19: {  # משק אורגני בן יהודה — mypips.app/organicfarm
        "B": "משפחת בן יהודה",
        "H": "בית יצחק שער חפר",
        "I": "ירקות",
        "J": "כן",
        "K": "לא",
        "L": "איסוף עצמי במשק",
        "M": "כשר",
        "N": ("משק משפחתי בבית יצחק שמגדל מעל עשרים שנה מגוון פירות "
              "וירקות אורגניים בפרדסים, חממות ושטחים פתוחים, תוך שמירה "
              "על איכות המוצר ועל בריאות הסביבה. חבר בארגון לחקלאות "
              "אורגנית ומפוקח על ידי משרד החקלאות. פתוח לקהל בימי חמישי "
              "ושישי, הזמנות לאיסוף עצמי בכל ימות השבוע."),
        "O": "פירות וירקות אורגניים ממשק משפחתי ותיק",
        "Q": "ben-yehuda-organic",
        "U": "לא",
        "V": "כן",
    },
    20: {  # הענתיות — mypips.app/anatiyot
        "H": "כל הארץ",
        "I": "פירות",
        "J": "כן",
        "K": "כן",
        "L": "תל אביב, ירושלים, חיפה ומרכז הארץ",
        "N": ("הענתיות — רכישה קבוצתית של פירות וירקות אורגניים, עלים "
              "ירוקים ואגוזים ברמה הגבוהה ביותר, יחד עם מגוון מוצרים "
              "משלימים — הכל טרי, איכותי ו-100% טבעוני. קטיף טרי ישירות "
              "מחקלאים ברחבי הארץ."),
        "O": "פירות וירקות אורגניים בקטיף טרי מהחקלאי",
        "Q": "anatiyot",
        "U": "לא",
        "V": "כן",
    },
    21: {  # מהחקלאי לצרכן — mypips.app/guy.oz.topper
        "B": "גיא עוז טופר",
        "H": "כל הארץ",
        "I": "ירקות",
        "J": "כן",
        "K": "לא",
        "L": "נקודות איסוף ברחבי הארץ",
        "N": ("מהחקלאי לצרכן — חקלאות ותוצרת ישראלית, פירות וירקות "
              "ישירות מחקלאים מקומיים בלי מתווכים. שוק עודפים ונקודות "
              "איסוף ברחבי הארץ, בהזמנה מראש דרך mypips."),
        "O": "פירות וירקות ישירות מהחקלאי הישראלי",
        "Q": "guy-oz-topper",
        "U": "לא",
        "V": "לא",
    },
}


def is_empty(value):
    if value is None:
        return True
    if isinstance(value, str) and value.strip() == "":
        return True
    return False


def main():
    wb = load_workbook(SRC)
    ws = wb.active

    filled = 0
    skipped = 0
    rows_touched = 0

    for row, payload in ENRICHMENTS.items():
        row_changed = False
        for col_letter, new_value in payload.items():
            col_idx = COL[col_letter]
            cell = ws.cell(row=row, column=col_idx)
            if is_empty(cell.value):
                cell.value = new_value
                filled += 1
                row_changed = True
            else:
                skipped += 1
        if row_changed:
            rows_touched += 1

    wb.save(DST)

    print(f"\n=== Enrichment complete ===")
    print(f"  Output:        {DST.name}")
    print(f"  Rows touched:  {rows_touched}")
    print(f"  Cells filled:  {filled}")
    print(f"  Cells skipped (already had data): {skipped}")


if __name__ == "__main__":
    main()
