import logging
from datetime import datetime, timedelta, timezone
from html import escape as _html_escape
from typing import NamedTuple
from uuid import UUID

import httpx
from fastapi import (
    APIRouter,
    Body,
    Depends,
    File,
    HTTPException,
    Query,
    Request,
    UploadFile,
)
from sqlalchemy import func, text
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import Session, joinedload, selectinload

from app.auth import require_admin
from app.utils.clock import business_days_waiting, israel_now, israel_today
from app.utils.pii import mask_phone
from app.config import settings
from app.services.auth_notifications import (
    notify_producer_approved,
    notify_producer_changes_requested,
)
from app.services.delivery_validation import (
    ensure_exclusion_requires_nationwide,
    ensure_nationwide_requires_delivery,
)
from app.services.email import send_email

# MEH-1242: SITE_DOMAIN is the single canonical domain constant. Import it —
# never re-declare the literal.
from app.services.onboarding_followup import SITE_DOMAIN
from app.services.whatsapp import send_text
from app.database import get_db
from app.rate_limit import limiter
from app.models import (
    DeliveryArea,
    HomeProduct,
    Producer,
    ProducerCategory,
    Product,
    User,
)
from app.schemas.schemas import (
    GrantVerifiedIn,
    LicenseExpiryReminderOut,
    LicenseExpiryReminderRow,
    ProducerAdminCreate,
    ProducerAdminOut,
    ProducerRejectIn,
    ProducerAdminUpdate,
    RejectionPresetOut,
    RequestChangesIn,
    StoryCardUploadRequest,
)
from app.services.license_validation import (
    categories_require_license,
    ensure_license_for_categories,
)
from app.services.producer_queries import (
    attach_badge_fields,
    create_primary_branch_location,
    upsert_primary_branch_location,
)
from app.slug_utils import is_reserved, rejected_characters

# MEH-2020 — `_slugify` used to be a full copy of `slug_utils.slugify` here,
# and a third copy lived in `producer_import.py`. Three generators of a public
# URL is three chances to drift, and MEH-2021 already caught one docstring
# describing all three wrongly. The name stays so every call site and the
# MEH-2021 corpus keep working; only the body is gone.
from app.slug_utils import slugify as _slugify
from app.utils.sql import LIKE_ESCAPE, escape_like

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/admin", tags=["admin"])


def _guard_supplied_slug(raw: str) -> None:
    """Validate a slug an admin typed, before it is normalised into one.

    MEH-2020. Two rejections, and the order between them is the ruling's, not a
    detail: **reserved is asked first, of the raw value under NFKC + casefold.**
    Ask it of `_slugify(raw)` instead and the fullwidth "ａbout" reports as a
    charset error, because the stripping already removed the character that made
    it reserved — the same word, refused for the wrong reason and with an error
    message that sends the admin to fix the wrong thing.

    The charset rejection is a 422 carrying a code plus the offending characters,
    not a hard-coded Hebrew sentence (MEH-1943). The reserved rejection keeps its
    existing 400 and its existing Hebrew string: changing a live contract that
    nothing in this ticket asked about is scope this change does not own.
    """
    if is_reserved(raw):
        raise HTTPException(
            status_code=400, detail="שם זה שמור לשימוש האתר. בחרי שם אחר."
        )
    bad = rejected_characters(raw)
    if bad:
        raise HTTPException(
            status_code=422,
            detail={"code": "slug_charset_not_allowed", "characters": bad},
        )


def _yes_no(value) -> bool:
    """Parse Hebrew/English yes/no values from Excel."""
    if value is None:
        return False
    if isinstance(value, bool):
        return value
    s = str(value).strip().lower()
    return s in ("כן", "yes", "y", "true", "1", "v", "✓")


def _ensure_unique_slug(
    db: Session, base_slug: str, exclude_id: UUID | None = None
) -> str:
    """Append -2, -3, ... until slug is unique and not reserved."""
    if not base_slug:
        return base_slug
    candidate = base_slug
    counter = 2
    while True:
        if not is_reserved(candidate):
            q = db.query(Producer).filter(Producer.slug == candidate)
            if exclude_id:
                q = q.filter(Producer.id != exclude_id)
            if not q.first():
                return candidate
        candidate = f"{base_slug}-{counter}"
        counter += 1


def _mint_slug_if_absent(db: Session, producer: Producer) -> None:
    """MEH-1817 — publish-time canonicalization of the producer's URL.

    Self-registration (`auth.py`, both branches) builds `Producer(...)` with no
    slug, and `_slugify` ran only on the admin-create / admin-update / import
    paths — so a self-registered business stayed `slug=NULL` forever and its
    public page fell back to `/producer/{uuid}`, losing the Hebrew URL the
    catalog's SEO is built on.

    Approval rather than registration: the name can still change during review,
    and a slug minted from a name that then changes is worse than none — it is
    a wrong URL that has to be redirected.

    Guarded on `not producer.slug`, so re-approving a business never rewrites
    an existing (possibly admin-chosen) slug.

    `or None` is load-bearing: `_ensure_unique_slug` returns "" unchanged for an
    empty base, so a name that slugifies to nothing must leave the column NULL.
    An empty string is neither a slug nor NULL — it breaks the id-URL fallback's
    NULL check AND satisfies the `not producer.slug` guard forever after, so no
    later approval could repair it.
    """
    if not producer.slug:
        producer.slug = _ensure_unique_slug(db, _slugify(producer.name)) or None


def _apply_approval_state(producer: Producer) -> None:
    """The complete set of column writes that "approved" means.

    Single owner, because `_persist_approval`'s retry path must re-apply every
    one of them after `db.rollback()` discards the first attempt. Two copies of
    this list would drift the moment a fourth field joins approval: the handler
    would set it, the rollback would discard it, and the retry would commit a
    row missing it — silently, and only on the collision path, which is the one
    path no ordinary test exercises.

    The slug is deliberately NOT here. It is minted by `_mint_slug_if_absent`,
    which needs the session (it queries for collisions) and must run *after*
    the re-read on the retry path so it sees the winner's committed slug.
    """
    producer.status = "approved"
    # MEH-1011: clear the request-changes trail on approve — the completion
    # request (if any) is resolved once the business is approved.
    producer.requested_changes = None
    producer.changes_requested_at = None
    # MEH-2210: symmetric clearing of the rejection trail — an approved
    # business must not carry a stale "לא אושרה" reason into its dashboard
    # (the banner reads producer_rejection_reason off GET /auth/me).
    # `resubmission_count` is deliberately NOT reset: it is history, and the
    # cap is per business, not per rejection.
    producer.rejection_reason = None
    producer.rejection_reason_code = None


class ApprovalOverrides(NamedTuple):
    """The admin's explicit "yes, I have checked this out-of-band" flags.

    MEH-2121. Bundled rather than threaded as two more booleans because
    `_persist_approval` already takes five arguments and `max-args = 5` is
    enforced on this file (`backend/pyproject.toml:116`; admin.py carries no
    per-file-ignore, unlike producers.py / auth.py). A sixth positional would
    have to be paid for by widening that ignore list, which is the wrong
    direction — so the bundle keeps both signatures at their current width and
    a third override, if one ever arrives, costs nothing.

    Both default False: the guards are ON unless an admin explicitly asks for
    the door, which is the MEH-971 property this mirrors.
    """

    without_license: bool = False
    unverified_phone: bool = False


def _assert_approvable(
    db: Session, producer: Producer, overrides: ApprovalOverrides, admin_id
) -> None:
    """The approval gates, in one owner, run against the row being written.

    MEH-2017. These lived inline in `approve_producer` and therefore ran only
    against the object fetched at the top of that handler. `_persist_approval`'s
    retry path re-reads the row after a rollback, so a producer whose photos or
    license number were cleared inside that window could be approved by the
    retry although the main path would have rejected it — and `_persist_approval`
    was the only place in the codebase able to approve without passing a gate.

    One function called twice, deliberately not a validator layer or a
    decorator: there are two call sites and no third is expected.

    The override warning lives HERE rather than in the handler so the retry path
    is audited too. A silent override on the one path this function exists to
    protect would be the same defect wearing a smaller diff.
    """
    # MEH-799: approval gate — a business never goes public without at least
    # one photo. Validation only (no schema change); registration/publish
    # flows untouched — the gate lives at the moment of approval.
    if not producer.images:
        raise HTTPException(
            status_code=422,
            detail="לא ניתן לאשר בית עסק ללא תמונה. בקשי מבעלת העסק להעלות תמונה אחת לפחות.",
        )
    # MEH-971 chunk 4: license-pending approval guard — the approve-gate mirror
    # of register-time ensure_license_for_categories. 422 matches the photo gate
    # above (not MEH-769's 409, which is for status transitions).
    category_ids = [c.id for c in producer.categories]
    license_missing = not (producer.producer_license_number or "").strip()
    needs_license = categories_require_license(db, category_ids) and license_missing
    if needs_license and not overrides.without_license:
        raise HTTPException(
            status_code=422,
            detail="לא ניתן לאשר בית עסק בקטגוריה הדורשת רישיון יצרן ללא מספר רישיון. אמתי את הרישיון, או אשרי עם דריסה מפורשת.",
        )
    if needs_license and overrides.without_license:
        # Audit trail: an admin bypassed the license guard — visible in Railway logs.
        logger.warning(
            "approve_producer: license-pending override for producer %s by admin %s",
            producer.id,
            admin_id,
        )

    # MEH-2121: the WhatsApp number is the channel every customer contact runs
    # through, so approving an unverified one publishes a page whose only CTA
    # may go nowhere. submission_gate makes the same argument for the owner's
    # submit gate; this is the approve-time mirror, for the older/bypassing
    # routes that reach approval without having passed it.
    #
    # 422, matching the photo and licence gates directly above — this is an
    # approve-time CONTENT gate, and this file's split is 422 for those, 409 for
    # status transitions (the draft blocks in the two handlers, MEH-769's
    # toggle guard, request_producer_changes').
    #
    # RULED by Sapir 18/08, and recorded because the ticket said otherwise: the
    # MEH-2121 AC specified 409 here. That was a spec error, caught
    # independently by this PR's Phase 0 and by the CI reviewer, and corrected
    # before merge rather than after a client had branched on it. The card's
    # description now matches. Do not "restore" 409.
    if not producer.phone_verified and not overrides.unverified_phone:
        raise HTTPException(
            status_code=422,
            detail="לא ניתן לאשר בית עסק ללא אימות מספר הוואטסאפ. בקשי מבעלת העסק לאמת את המספר, או אשרי עם דריסה מפורשת.",
        )
    if not producer.phone_verified and overrides.unverified_phone:
        # Same audit shape as the license override directly above — a log line,
        # deliberately, because that is the mechanism this repo already uses.
        logger.warning(
            "approve_producer: unverified-phone override for producer %s by admin %s",
            producer.id,
            admin_id,
        )


SLUG_UNIQUE_CONSTRAINT = "producers_slug_key"


def _is_slug_collision(exc: IntegrityError) -> bool:
    """True only for a `producers.slug` unique violation.

    `db.commit()` flushes the WHOLE session, so a bare `except IntegrityError`
    around it catches any constraint violation on any pending row — not just
    the one this function is recovering from. Retrying an unrelated violation
    is the dangerous branch: the rollback discards it, the retry re-applies
    only the approval columns, and the request commits a partially-applied
    transaction and returns 200. A wrong row, silently, on the path nothing
    exercises.

    So the default is to **re-raise**. `producers_slug_key` is the only unique
    constraint on `producers` today (verified against `pg_constraint`), but
    that is a fact about today's schema, not an invariant — this check is what
    makes a future second unique index surface as a 500 with its real cause
    instead of being swallowed by a retry meant for slugs.

    Falls back to the message text when the driver exposes no `diag` (psycopg2
    populates it; SQLite and some wrappers do not), and re-raises when neither
    source names the constraint. Unknown means raise, never retry.
    """
    orig = getattr(exc, "orig", None)
    constraint = getattr(getattr(orig, "diag", None), "constraint_name", None)
    if constraint:
        return constraint == SLUG_UNIQUE_CONSTRAINT
    return SLUG_UNIQUE_CONSTRAINT in str(orig or exc)


def _persist_approval(
    db: Session,
    producer_id: UUID,
    producer: Producer,
    overrides: ApprovalOverrides,
    admin_id,
) -> Producer:
    """Mint the slug and commit, retrying once on a unique-slug collision.

    `producers.slug` is UNIQUE (`models/models.py:112`) and `_ensure_unique_slug`
    is SELECT-then-return with no lock, so two admins approving two same-named
    businesses in one window both see the candidate free and both take it. The
    second commit then violates the constraint.

    That exposure is NEW and belongs to MEH-1817: before the mint, this handler
    never wrote `slug`, so it could not reach the constraint at all. Left
    uncaught it would be a raw 500 **and** roll back the entire transaction —
    the approval, the status flip, the requested-changes clear — so the losing
    admin's approval would fail silently with no notification sent.

    Recovery follows the house pattern (`reviews.py:288`, `reports.py:62`,
    `favorites.py:72`, `referrals.py:54`, `producer_me.py:228`): roll back and
    retry once. The retry re-reads the row, so `_ensure_unique_slug` now sees
    the winner's committed slug and suffixes past it.

    Not retried twice. A second collision would mean a third admin approving
    the same name in the same instant; letting that raise is more honest than
    looping, and a 500 whose cause is one frame up beats one nobody can explain.

    Only a `producers_slug_key` violation is retried — see `_is_slug_collision`.
    """
    _mint_slug_if_absent(db, producer)
    try:
        db.commit()
        return producer
    except IntegrityError as exc:
        db.rollback()
        if not _is_slug_collision(exc):
            raise
    producer = db.query(Producer).filter(Producer.id == producer_id).first()
    if not producer:
        raise HTTPException(status_code=404, detail="בית עסק לא נמצא") from None
    # MEH-2017: the gates run on BOTH paths. This row is fresh from the database
    # and is the one about to be committed, so it is re-validated rather than
    # trusted from the pre-collision fetch — if another transaction stripped the
    # images or the license number inside the rollback window, this raises 422
    # exactly as the main path would have. One owner, called twice.
    _assert_approvable(db, producer, overrides, admin_id)
    _apply_approval_state(producer)
    _mint_slug_if_absent(db, producer)
    db.commit()
    return producer


def _apply_categories(db: Session, producer: Producer, category_ids: list[int]):
    db.query(ProducerCategory).filter(
        ProducerCategory.producer_id == producer.id
    ).delete()
    # MEH-1297: payload order = stored order (position 0 = primary).
    for pos, cid in enumerate(category_ids):
        db.add(ProducerCategory(producer_id=producer.id, category_id=cid, position=pos))


def _apply_delivery_cities(db: Session, producer: Producer, cities: list[str]):
    db.query(DeliveryArea).filter(DeliveryArea.producer_id == producer.id).delete()
    for city in cities:
        city = (city or "").strip()
        if not city:
            continue
        db.add(DeliveryArea(producer_id=producer.id, city=city))


def _attach_aging(producer) -> None:
    """MEH-2110 — hydrate `business_days_waiting` for an admin list row.

    The server is the single source of truth for aging: one implementation,
    covered by tests/test_admin_queue_sla_aging.py, instead of a date
    calculation duplicated per client and drifting.

    A draft measures from `created_at` because it has no submission to measure
    from. Everything else uses the same COALESCE the queue ordering uses, so
    the badge and the sort can never disagree about which row is oldest.

    Called by BOTH admin list routes. `ProducerAdminOut` declares the field, so
    a route that returned rows without calling this would serialise a silent
    `0` on every row — a number that looks measured and is not.
    """
    producer.business_days_waiting = business_days_waiting(
        producer.created_at
        if producer.status == "draft"
        else (producer.submitted_for_review_at or producer.created_at)
    )


# MEH-1494 chunk B: the annual-review predicate, extracted so the filter reads
# as one named thing in `list_producers` and can be asserted directly.
#
# The window is the TripAdvisor/Michelin pattern the card cites: an editorial
# pick carries a visible clock and is re-examined, rather than being permanent.
#
# `recommended_at IS NULL` is IN the due set on purpose, and it is the case
# that matters most today. Chunk A deliberately did not backfill a date onto
# rows picked before the column existed, because inventing one would fabricate
# a decision date; NULL means "picked before there was a clock — review it now",
# which is exactly how Michelin treats a star that has not been re-examined.
#
# 365 days rather than a calendar year: the difference is a leap day on a
# review cadence measured in months, and `timedelta` costs no dependency. If
# the boundary ever has to be exact, this is the one line to change.
REVIEW_WINDOW_DAYS = 365


def _recommended_review_due_clause():
    """SQL: the row carries the pick AND its clock is unset or past the window."""
    cutoff = israel_now() - timedelta(days=REVIEW_WINDOW_DAYS)
    return Producer.is_recommended.is_(True) & (
        Producer.recommended_at.is_(None) | (Producer.recommended_at < cutoff)
    )


@router.get("/producers", response_model=list[ProducerAdminOut])
def list_producers(
    status: str | None = Query(
        None,
        pattern="^(draft|pending|approved|rejected|inactive|all)$",
    ),
    search: str | None = None,
    # MEH-1494 chunk B: the annual-review view. Default False so every existing
    # caller — the admin toolbar included — sees exactly the list it saw before.
    recommended_review_due: bool = Query(False),
    user: User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    q = db.query(Producer).options(
        joinedload(Producer.categories),
        joinedload(Producer.products),
        joinedload(Producer.delivery_areas),
        # MEH-2060: needed by attach_badge_fields' pickup_points/offers_pickup
        # derivation below. selectinload (not joinedload) — a 4th collection
        # joinedload here would widen the existing 3-way cartesian.
        selectinload(Producer.locations),
    )
    # MEH-2100: three-way, and the DEFAULT (no `status` param) is the one that
    # changed. The admin toolbar sends no param for its "כל הסטטוסים" option
    # (use-admin-producers.js omits it when the selection is "all"), so that
    # default IS the review queue the admin actually looks at — and drafts must
    # not be in it. A draft is a business that has not asked to be reviewed;
    # showing it is the queue-noise this ticket exists to remove.
    #
    #   no param      -> everything EXCEPT draft  (the working queue)
    #   ?status=all   -> genuinely everything, drafts included (the escape
    #                    hatch — "all" must not quietly mean "all but one")
    #   ?status=draft -> drafts only (the new "טיוטות" option, visibility only)
    #
    # `?status=pending` used to group two values, pending + pending_whatsapp;
    # the second was removed in MEH-2124 along with the state itself, so the
    # filter is a plain equality now and the branch that special-cased it is
    # gone.
    if status == "all":
        pass  # no status filter at all — the only view that shows drafts
    elif status:
        q = q.filter(Producer.status == status)
    else:
        q = q.filter(Producer.status != "draft")
    if search:
        # F13 (MEH-1188): escape LIKE metacharacters so a user-supplied % / _
        # matches literally instead of acting as a wildcard (same class as F1).
        like = f"%{escape_like(search)}%"
        q = q.filter(
            Producer.name.ilike(like, escape=LIKE_ESCAPE)
            | Producer.city.ilike(like, escape=LIKE_ESCAPE)
        )
    if recommended_review_due:
        q = q.filter(_recommended_review_due_clause())
    # MEH-2110: the review queue is worked oldest-first, because the "עד 3 ימי
    # עסקים" promise starts at submission and a business that has waited longest
    # is the one closest to breaking it. Newest-first (the old default) let an
    # old row sink under fresh arrivals and the promise break silently.
    #
    # WHICH VIEWS SORT ASC, and why it is not just `?status=pending`: the
    # admin's working view is the DEFAULT (no param), which this file's filter
    # above resolves to `status != "draft"` — NOT to the pending filter.
    # The ticket's wording assumed those were the same view; they are not, so
    # sorting only the explicit pending filter would have left the screen the
    # admin actually opens completely unchanged. Explicit approved/rejected/
    # inactive/draft filters keep the existing newest-first order, where recency
    # is the useful axis and there is no SLA to track.
    queue_view = status is None or status == "pending"
    if queue_view:
        # COALESCE, not submitted_for_review_at alone: rows seeded before
        # MEH-2100 (and every draft) have a NULL stamp, and a bare column sort
        # would bunch them all at one end regardless of real age.
        order_key = func.coalesce(
            Producer.submitted_for_review_at, Producer.created_at
        ).asc()
    else:
        order_key = Producer.created_at.desc()
    results = q.order_by(order_key).all()
    # MEH-2060: this endpoint never called attach_badge_fields before, so
    # pickup_points (and every other computed badge field) silently defaulted
    # to Pydantic's False/0 in this response — a pre-existing gap, not
    # something this change introduces. Calling it here is what makes
    # AdminProducersTable.jsx's pickup badge (:128) agree with every other
    # consumer surface instead of reading a raw, no-longer-authoritative column.
    for p in results:
        attach_badge_fields(p)
        _attach_aging(p)
    return results


@router.post("/producers", response_model=ProducerAdminOut, status_code=201)
def admin_create_producer(
    data: ProducerAdminCreate,
    user: User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    """Admin-created producers are auto-approved."""
    slug = data.slug or _slugify(data.name)
    # Reject explicit reserved slugs; auto-generated slugs get suffixed by _ensure_unique_slug.
    if data.slug:
        _guard_supplied_slug(data.slug)
    slug = _ensure_unique_slug(db, slug)

    # MEH-530: same conditional-required guard as the public endpoints —
    # admin form can still persist non-regex license values verbatim
    # (manual-approval flow), but missing-license-for-required-category
    # is still a 422.
    ensure_license_for_categories(db, data.category_ids, data.producer_license_number)

    producer = Producer(
        name=data.name,
        contact_name=data.contact_name,
        description=data.description,
        short_description=data.short_description,
        city=data.city,
        lat=data.lat,
        lng=data.lng,
        phone=data.phone,
        instagram=data.instagram,
        website=data.website,
        whatsapp_group=data.whatsapp_group,
        # MEH-17
        primary_contact_method=data.primary_contact_method or "whatsapp",
        contact_email=data.contact_email,
        # MEH-296 3d: admin-create parity for the new channels.
        facebook=data.facebook,
        external_order_form=data.external_order_form,
        slug=slug,
        top_product_name=data.top_product_name,
        price_range=data.price_range,
        grass_fed=data.grass_fed,
        organic_certified=data.organic_certified,
        has_delivery=data.has_delivery,
        # MEH-2060: pickup_points stopped being written here — it's derived
        # from producer_locations rows now (attach_badge_fields below), and a
        # freshly-created producer has none yet regardless of this form field.
        kosher=data.kosher,
        producer_license_number=data.producer_license_number,
        admin_notes=data.admin_notes,
        # MEH-766 ch3: is_verified no longer set on admin create — column default
        # False applies; verification is via grant-verified (verified_at) only.
        images=data.images or [],
        # MEH-213 — location mode
        has_physical_location=data.has_physical_location,
        offers_delivery=data.offers_delivery,
        delivery_nationwide=data.delivery_nationwide,
        # MEH-1255: nationwide exclusion list — schema validator already
        # rejected excluded-without-nationwide on create.
        delivery_excluded_cities=data.delivery_excluded_cities,
        # MEH-903 A: the legacy delivery_cities column is no longer written —
        # delivery_areas (via _apply_delivery_cities below) is the single store.
        # Column stays declared (drop = Chunk C); new rows keep the [] default.
        status="approved",  # admin = pre-approved
    )
    db.add(producer)
    db.flush()

    _apply_categories(db, producer, data.category_ids)
    _apply_delivery_cities(db, producer, data.delivery_area_cities)
    # MEH-1921: this is the FIFTH create-from-payload path, and the one easiest
    # to miss — `_apply_delivery_cities` is shared with the PUT route (:286),
    # so classifying it by call site reads "edit path, leave alone" and hides
    # that THIS caller creates. Like the importer, the row is born
    # `status="approved"` and is live immediately.
    #
    # Unlike the four signup paths, `ProducerAdminCreate` DOES carry
    # `offers_delivery` (schemas.py:1431), so an admin can state it — and an
    # explicit `false` alongside delivery cities is a deliberate declaration
    # that MEH-1848's predicates are right to exclude. Only the UNSTATED case
    # is derived, which `model_fields_set` distinguishes and a plain falsy
    # check cannot. Same distinction `_sync_active_offer` (producer_me.py)
    # already relies on: "omitted" and "explicitly false" are different answers.
    if data.delivery_area_cities and "offers_delivery" not in data.model_fields_set:
        producer.offers_delivery = True

    # MEH-2059 (MEH-1938 chunk 4b): the admin half of the dual-write. Same
    # helper the four signup paths call (producer_queries.py:351), so this is
    # the SIXTH create-from-payload site and the field contract cannot drift
    # between them. Declines by itself when a coordinate is missing.
    #
    # This site can only ever create: the producer was flushed six lines up, so
    # it owns no locations yet — `upsert_...` would find nothing and delegate
    # here anyway. Calling the create helper directly says that out loud.
    #
    # REUSES: backend/app/services/producer_queries.py:449 — the same call, in
    # the same position (after flush, before commit), on the public-create path.
    create_primary_branch_location(db, producer)

    db.commit()
    db.refresh(producer)
    attach_badge_fields(producer)
    return producer


# MEH-2072: how far ahead the licence reminder looks. One constant, not a query
# param — mirroring EXPIRY_REMINDER_WINDOW_DAYS in admin_kashrut.py:31 and for
# the same reason recorded there: a caller-chosen window lets a mistyped value
# sweep in every business on file, and the ticket scopes this to 30 days.
LICENSE_EXPIRY_REMINDER_WINDOW_DAYS = 30


@router.get("/license-expiry-reminders", response_model=LicenseExpiryReminderOut)
@limiter.limit("60/minute")
def license_expiry_reminders(
    request: Request,
    user: User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    """MEH-2072: approved businesses whose licence expires within 30 days.

    Read-only. Nothing is sent, nothing is hidden, nothing is un-verified — v1
    policy is capture + remind, and the admin pulls this list herself. That is
    why this is a GET with no `dry_run` flag, unlike its kashrut sibling
    (admin_kashrut.py:154) which POSTs because it dispatches WhatsApp. Push is a
    future ticket.

    Selection, and every clause is load-bearing:

    * `license_expires_at IS NOT NULL` — NULL means "not captured yet", never
      "no expiry". Without this clause every producer predating MEH-2072 would
      appear in the reminder list forever, which is the failure that would get
      the whole feature switched off.
    * `>= israel_today()` — an already-lapsed licence is deliberately EXCLUDED.
      This mirrors the kashrut endpoint's reasoning: the list exists to catch a
      licence before it lapses. A lapsed one is a different problem needing a
      different action (that is the future enforcement decision), and letting it
      sit here would grow an unactionable backlog that buries the live rows.
    * `<= horizon` — the 30-day window.
    * `status == "approved"` — a pending business is already in the review
      queue, where the admin sees the licence anyway.

    Israel's calendar day, not UTC's: `license_expires_at` is a DATE column, and
    `israel_today()` is what makes the comparison calendar-day against
    calendar-day with no zone arithmetic (see the migration docstring).

    `60/minute`, matching the read-only `list_kashrut_requests`
    (admin_kashrut.py:36) rather than the `10/hour` its expiry-reminder sibling
    carries — that one is throttled hard because it DISPATCHES WhatsApp, and a
    read has no outbound side effect to ration.

    Note this is the first `@limiter.limit` in this module: `admin.py` currently
    has none at all. That is a pre-existing gap, not a local convention to copy
    — `backend/app/routers/CLAUDE.md` states the per-route limiter stack is
    mandatory here, so the new route follows the documented rule rather than the
    surrounding silence. Retro-fitting the rest of the module is out of scope.
    """
    today = israel_today()
    horizon = today + timedelta(days=LICENSE_EXPIRY_REMINDER_WINDOW_DAYS)
    producers = (
        db.query(Producer)
        .filter(
            Producer.license_expires_at.isnot(None),
            Producer.license_expires_at >= today,
            Producer.license_expires_at <= horizon,
            Producer.status == "approved",
        )
        .order_by(Producer.license_expires_at.asc())
        .all()
    )

    rows = [
        LicenseExpiryReminderRow(
            producer_id=producer.id,
            name=producer.name,
            # mask_phone handles the missing-phone case; unlike the kashrut
            # endpoint there is no `phone IS NOT NULL` filter, because a
            # business with no phone on file still needs its licence chased —
            # the admin just reaches it another way. Filtering it out would hide
            # the row entirely, which is the opposite of what this list is for.
            phone_masked=mask_phone(producer.phone),
            expires_at=producer.license_expires_at,
            # Computed here, not per client: same reasoning as
            # business_days_waiting on ProducerAdminOut — the number the admin
            # reads and the order the rows arrive in must come from one clock.
            days_remaining=(producer.license_expires_at - today).days,
            producer_license_number=producer.producer_license_number,
        )
        for producer in producers
    ]

    return LicenseExpiryReminderOut(
        window_days=LICENSE_EXPIRY_REMINDER_WINDOW_DAYS,
        total=len(rows),
        rows=rows,
    )


def _apply_recommended_pick(producer: Producer, payload: dict) -> None:
    """Normalise the editorial pick and stamp/clear its date on a transition.

    MEH-1494 chunk B. The stamp is driven by the TRANSITION, never by the
    value: `recommended_at` is what makes the annual review possible (the
    TripAdvisor/Michelin pattern on the card), so a row re-saved while already
    picked keeps its original date. Stamping on the value instead would reset
    the review window on every unrelated admin edit to that producer — which is
    the whole clock, silently disarmed.

    Gated on the PAYLOAD, not on the resulting value, matching this handler's
    `exclude_unset` semantics throughout: an admin editing only the name never
    touches the stamp.

    `bool(...)` on both sides because `is_recommended` is nullable and an
    explicit `is_recommended: null` reaches here — a NULL flag is not a pick,
    so it lands on the un-pick branch.

    `recommended_note` is deliberately NOT cleared on an un-pick: it is the
    editor's record of a decision that was actually made, and erasing it would
    destroy the only trace of why — the thing ADR-030 requires to be
    defensible. The next pick overwrites it.
    """
    if "is_recommended" not in payload:
        return
    was = bool(producer.is_recommended)
    will_be = bool(payload["is_recommended"])
    # Normalise BEFORE the setattr loop reaches the column. `is_recommended` is
    # nullable while `ProducerAdminOut.is_recommended` is a plain `bool`, so a
    # PUT carrying an explicit `is_recommended: null` wrote NULL and then died
    # in response validation — a 500 on a successful write. Measured 06/09
    # against the pre-chunk-B code, so this is a latent defect this chunk
    # surfaced rather than one it introduced; the column stays nullable
    # (narrowing it is Alembic, and this handler is not where that belongs).
    payload["is_recommended"] = will_be
    if will_be and not was:
        producer.recommended_at = israel_now()
    elif was and not will_be:
        # Clearing the date takes the row OUT of the review list rather than
        # leaving it there forever as permanently overdue.
        producer.recommended_at = None


@router.put("/producers/{producer_id}", response_model=ProducerAdminOut)
def admin_update_producer(
    producer_id: UUID,
    # MEH-1287 chunk B: the ADMIN shape — ProducerUpdate plus the editor-only
    # fields. The owner PUT (producer_me.py) keeps ProducerUpdate, which is what
    # keeps `in_season_until` off the owner's schema entirely rather than merely
    # filtered out of it.
    data: ProducerAdminUpdate,
    user: User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    producer = db.query(Producer).filter(Producer.id == producer_id).first()
    if not producer:
        raise HTTPException(status_code=404, detail="בית עסק לא נמצא")

    payload = data.model_dump(exclude_unset=True)
    category_ids = payload.pop("category_ids", None)
    delivery_area_cities = payload.pop("delivery_area_cities", None)
    # MEH-903 A: stop writing the legacy delivery_cities column — delivery_areas
    # is the single store now. Pop it out of the payload so the bulk setattr loop
    # below can't resurrect the write. Column stays declared (drop = Chunk C).
    payload.pop("delivery_cities", None)
    # MEH-2060: same no-op-pop pattern, same reason — pickup_points is derived
    # from producer_locations rows now (attach_badge_fields below); ProducerForm
    # still has a checkbox that sends this field (frontend out of scope for this
    # ticket, flagged separately), but the bulk setattr loop must not write it.
    payload.pop("pickup_points", None)
    # MEH-1255: effective-state guard — excluded cities require nationwide.
    ensure_exclusion_requires_nationwide(producer, payload)
    # MEH-1879: same shape — nationwide delivery requires the delivery flag,
    # or the DB CHECK (MEH-1849) turns a partial update into a 500.
    ensure_nationwide_requires_delivery(producer, payload)

    # MEH-530: PATCH semantics — guard against the EFFECTIVE state after
    # the update. If category_ids is being changed → use the new list,
    # otherwise read existing producer-category join rows. Same for license:
    # if the field is in the payload (even explicitly None to clear) → use
    # that, otherwise keep the current column. Helper short-circuits to OK
    # when no required-category is touched, so this is cheap on non-license
    # admin edits.
    effective_category_ids = (
        category_ids
        if category_ids is not None
        else [c.id for c in producer.categories]
    )
    effective_license = (
        payload.get("producer_license_number")
        if "producer_license_number" in payload
        else producer.producer_license_number
    )
    ensure_license_for_categories(db, effective_category_ids, effective_license)

    # Keep slug unique if changed; reject reserved slugs.
    if "slug" in payload and payload["slug"]:
        _guard_supplied_slug(payload["slug"])
        candidate = _slugify(payload["slug"])
        payload["slug"] = _ensure_unique_slug(db, candidate, exclude_id=producer.id)

    # MEH-375: snapshot gallery BEFORE bulk setattr so we can diff and
    # destroy URLs the admin dropped AFTER db.commit succeeds. Order
    # matters — destroying before commit would orphan-leak in reverse
    # (assets gone, DB still references them) on a commit raise.
    old_images = list(producer.images or [])

    # MEH-1494 chunk B: the editorial clock, stamped BEFORE the setattr loop
    # below — that loop is what turns `producer.is_recommended` into the new
    # value, so the transition has to be read while it is still the old one.
    _apply_recommended_pick(producer, payload)

    for field, value in payload.items():
        setattr(producer, field, value)

    if category_ids is not None:
        _apply_categories(db, producer, category_ids)
    if delivery_area_cities is not None:
        _apply_delivery_cities(db, producer, delivery_area_cities)

    # MEH-2059 (MEH-1938 chunk 4b): the admin EDIT half of the dual-write.
    # Gated on the payload rather than on the resulting values, because
    # `exclude_unset=True` above means an admin editing only the name must not
    # re-mirror `Producer.address` over a location row the OWNER authored in
    # the dashboard. An explicit `lat: null` IS in the payload, so clearing the
    # coordinates still reaches the helper — which mirrors the clear instead of
    # leaving a stale point behind.
    #
    # Runs AFTER the setattr loop on purpose: the helper reads the producer
    # instance, so it must see the post-update values. That is the same reason
    # MEH-1939 takes the flushed Producer rather than five loose fields
    # (producer_queries.py:378-383) — the row is a mirror, and reading the
    # mirror's source off the instance is what makes them unable to drift.
    if "lat" in payload or "lng" in payload:
        upsert_primary_branch_location(db, producer)

    db.commit()
    db.refresh(producer)

    # MEH-375: post-commit cleanup. Helper handles set diff + dedup +
    # fail-open per-URL destroy.
    if "images" in payload:
        from app.cloudinary_utils import destroy_removed_images

        destroy_removed_images(
            old_images,
            producer.images or [],
            context="admin.admin_update_producer images",
        )

    attach_badge_fields(producer)
    return producer


# MEH-769 (HOT-002): the toggle is purely the visibility switch for an
# already-decided business — approved ⇄ inactive only. Any other source
# status (draft / pending / rejected) must go through the real
# approve_producer flow, which fires the MEH-509 side-effects (approval
# email, producer_approved_v1 WhatsApp, admin WhatsApp). Before this guard
# the bare `else` branch silently force-approved a REJECTED producer onto
# the public map, skipping every validation and notification.
_TOGGLEABLE_STATUSES = {"approved", "inactive"}


@router.post("/producers/{producer_id}/toggle-status")
def toggle_producer_status(
    producer_id: UUID,
    user: User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    """Toggle approved <-> inactive (hides from public listings).

    Refuses any other source status with 409 — use the approve/reject flow.
    """
    producer = db.query(Producer).filter(Producer.id == producer_id).first()
    if not producer:
        raise HTTPException(status_code=404, detail="בית עסק לא נמצא")
    # MEH-769: block the rejected/pending → approved force-flip. Final user-
    # facing Hebrew lives in the frontend message key
    # (admin.producers.toggle.invalid_transition); this detail is the API
    # contract / fallback.
    if producer.status not in _TOGGLEABLE_STATUSES:
        raise HTTPException(
            status_code=409,
            detail="לא ניתן לשנות סטטוס במצב הנוכחי — יש לאשר או לדחות את העסק דרך מסלול האישור.",
        )
    producer.status = "inactive" if producer.status == "approved" else "approved"
    db.commit()
    return {"detail": "Status toggled", "status": producer.status}


@router.delete("/producers/{producer_id}")
def admin_delete_producer(
    producer_id: UUID,
    user: User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    producer = db.query(Producer).filter(Producer.id == producer_id).first()
    if not producer:
        raise HTTPException(status_code=404, detail="בית עסק לא נמצא")

    # MEH-375 + MEH-510: capture every Cloudinary URL owned by this
    # producer BEFORE db.delete — the cascade detaches the relationship
    # and producer.images / Product.image_url / story_card_url become
    # unreachable after commit. Destroy runs AFTER commit so a
    # constraint / deadlock failure doesn't leave Cloudinary and DB
    # out of sync.
    #
    # MEH-510: story_card_url IS captured here. The reserved namespace
    # (mehamakor/producers/*) is protected by destroy_image's reject
    # list to keep the cleanup script from sweeping live story-cards,
    # but the producer-delete path is the one legitimate caller that
    # should free the slot — pass `bypass_reserved=True` to opt out.
    old_image_urls = list(producer.images or [])
    products = db.query(Product).filter(Product.producer_id == producer.id).all()
    old_product_urls = [p.image_url for p in products if p.image_url]
    old_story_card_url = producer.story_card_url
    # MEH-1335: owner photo lives at mehamakor/owner/owner_{producer_id} —
    # NOT in RESERVED_PUBLIC_ID_PREFIXES (only mehamakor/producers/ is), so
    # its destroy below needs no bypass. Same orphan class as story_card:
    # overwrite=True on upload prevents duplicates, not delete-orphans.
    old_owner_photo_url = producer.owner_photo_url

    # MEH-747: unlink any user pointing at this producer BEFORE db.delete.
    # User.producer_id has no ondelete (models.py), so deleting the producer
    # while a self-registered owner still references it violates
    # users_producer_id_fkey → 500. Mirrors the auth.py::delete_account fix.
    # is_producer is also cleared: the producer is permanently gone, so the
    # "durable flag" no longer reflects reality, and leaving it True would
    # lock the owner out of re-registering (409 at auth.py — MEH-669 family).
    # role is reset to "consumer" for the same consistency reason:
    # require_producer (auth.py:268-273) gates on role ALONE, so a leftover
    # role="producer" with producer_id=NULL is an orphan that passes the dep
    # then 404s on every /producers/me* handler (producer_me.py:75-76).
    # Resetting role keeps the (role, producer_id) pair consistent — mirrors
    # the atomic set in the register flow (auth.py:511-514).
    # Admin-created producers have no linked user → update is a no-op.
    db.query(User).filter(User.producer_id == producer.id).update(
        {"producer_id": None, "is_producer": False, "role": "consumer"},
        synchronize_session=False,
    )
    db.flush()

    # MEH-816: phone_otp_tokens cascade via the DB FK (ondelete=CASCADE) plus
    # passive_deletes=True on the Producer.otp_tokens backref (MEH-773 Chunk B),
    # so no explicit pre-delete is needed. DO NOT re-add one — passive_deletes
    # already pre-empts the NotNullViolation the old MEH-755 bulk-delete guarded.
    db.delete(producer)
    db.commit()

    # Post-commit orphan cleanup, fail-open per destroy_image contract.
    from app.cloudinary_utils import destroy_image

    for url in old_image_urls:
        destroy_image(url, context="admin.admin_delete_producer images")
    for url in old_product_urls:
        destroy_image(url, context="admin.admin_delete_producer product_image")
    # MEH-510: bypass_reserved=True — the producer is gone, the slot is now an orphan.
    destroy_image(
        old_story_card_url,
        bypass_reserved=True,
        context="admin.admin_delete_producer story_card",
    )
    # MEH-1335: owner photo — non-reserved namespace, default reject list OK.
    destroy_image(
        old_owner_photo_url,
        context="admin.admin_delete_producer owner_photo",
    )

    return {"detail": "Producer deleted"}


@router.post("/producers/import")
async def import_producers_excel(
    file: UploadFile = File(...),
    dry_run: bool = Query(
        True, description="Preview only — set false to actually save"
    ),
    user: User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    """Upload Excel/CSV file and import producers. dry_run=true returns preview only."""
    from io import BytesIO

    try:
        from openpyxl import load_workbook
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed")

    contents = await file.read()
    if len(contents) > 10_000_000:
        raise HTTPException(status_code=413, detail="קובץ גדול מדי — מקסימום 10MB")
    try:
        wb = load_workbook(BytesIO(contents), data_only=True)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"לא ניתן לקרוא את הקובץ: {e}")

    ws = wb.active
    # Skip header row
    rows = [list(r) for r in ws.iter_rows(min_row=2, values_only=True)]

    from app.services.producer_import import import_rows

    return import_rows(db, rows, dry_run=dry_run)


@router.get("/producers/pending", response_model=list[ProducerAdminOut])
def pending_producers(
    user: User = Depends(require_admin), db: Session = Depends(get_db)
):
    results = (
        db.query(Producer)
        .options(
            joinedload(Producer.categories),
            joinedload(Producer.products),
            joinedload(Producer.delivery_areas),
            # MEH-2060: see list_producers above — same derivation, same reason.
            selectinload(Producer.locations),
        )
        .filter(Producer.status == "pending")
        # NOTE (MEH-2110): deliberately still newest-first. This route has no
        # frontend consumer today, and re-ordering it is a behaviour change the
        # ticket does not ask for. The aging field below IS attached, because
        # ProducerAdminOut declares it and serialising a structural `0` here
        # would misreport rather than merely omit.
        .order_by(Producer.created_at.desc())
        .all()
    )
    for p in results:
        attach_badge_fields(p)
        _attach_aging(p)
    return results


@router.post("/producers/{producer_id}/approve")
def approve_producer(
    producer_id: UUID,
    # MEH-971 chunk 4: explicit admin override for the license-pending guard
    # below. Defaults False so the guard is on by default; an admin who has
    # verified a license out-of-band passes ?allow_without_license=true.
    allow_without_license: bool = Query(default=False),
    # MEH-2121: the same shape, for the WhatsApp-verification gate. Default
    # False, so an admin has to ask for the door in the URL.
    allow_unverified_phone: bool = Query(default=False),
    user: User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    producer = db.query(Producer).filter(Producer.id == producer_id).first()
    if not producer:
        raise HTTPException(status_code=404, detail="בית עסק לא נמצא")
    # MEH-2121: a draft never asked to be reviewed. Approving one publishes a
    # business that skipped the whole machine and leaves
    # `submitted_for_review_at` NULL, which is the column MEH-2110's SLA badge
    # counts from — so the queue would carry a row it cannot age. No override:
    # unlike the license and phone gates below, there is no legitimate reason to
    # approve something that was never submitted; the owner presses the button.
    #
    # Handler-only, and that is sufficient rather than an oversight: the retry
    # path re-reads the row, but nothing in the codebase moves an existing
    # producer BACK to draft — all three `status="draft"` writes are at creation
    # (auth.py:561, auth.py:768, producer_queries.py:542). A row that was not
    # draft when this line ran cannot be draft when the retry reads it.
    if producer.status == "draft":
        raise HTTPException(
            status_code=409,
            detail="העסק עדיין בטיוטה — טרם נשלח לבדיקה",
        )
    overrides = ApprovalOverrides(
        without_license=allow_without_license,
        unverified_phone=allow_unverified_phone,
    )
    # MEH-2017: the approval gates (MEH-799 photo, MEH-971 license, MEH-2121
    # phone) live in _assert_approvable, which _persist_approval also calls
    # after its re-read so the retry path cannot approve a row this path would
    # have rejected.
    _assert_approvable(db, producer, overrides, user.id)
    # MEH-1817: the approval column writes live in _apply_approval_state so the
    # retry path below can re-apply exactly the same set after a rollback.
    _apply_approval_state(producer)
    # MEH-1817: slug mint + commit, extracted so this handler stays under the
    # C901 ceiling. See _persist_approval for why the commit is retried.
    producer = _persist_approval(db, producer_id, producer, overrides, user.id)

    # MEH-509 PR1: capture primitives before any post-commit work — ORM
    # attributes are safe here (no expire_on_commit configured on this
    # session), but capturing decouples the notify calls from the model.
    p_name = producer.name
    p_phone = producer.phone
    p_slug = producer.slug
    p_id = producer.id

    producer_user = db.query(User).filter(User.producer_id == producer.id).first()
    if producer_user:
        _send_notification_email(
            producer_user.email,
            # MEH-2113: the celebratory headline, in the one place it is TRUE.
            # It was rejected for the registration screen (MEH-2100 item 9)
            # precisely because onboarding was not complete there; at approval
            # it is — the business is live on the site. Sapir-approved copy,
            # verbatim (16/08) and untouched by MEH-2151 — that ticket restructures
            # the BODY (CTA links, HTML part), never this subject line.
            "ברוכים הבאים למהמקור",
            # MEH-2151: p_slug is captured above (MEH-1817 mints it during this
            # very approval), so both parts can carry the /p/{slug} link. A
            # producer with no slug still gets a well-formed mail — both
            # builders drop their view-page block whole.
            _producer_approved_body(p_name, p_slug),
            html=_producer_approved_html(p_name, p_slug),
        )

    # MEH-509 PR1: fire producer_approved_v1 WhatsApp template to the
    # producer. Fail-open at the service layer — any failure here must
    # NOT block the 200 response (approval already committed above).
    notify_producer_approved(p_name, p_phone, p_slug, p_id)

    # Notify admin via WhatsApp
    if settings.admin_whatsapp_to:
        _send_whatsapp(
            settings.admin_whatsapp_to,
            f'✅ העסק "{p_name}" אושר במהמקור.',
        )

    return {"detail": "Producer approved"}


# MEH-226: the five canonical rejection reasons. This dict is the SINGLE
# source of truth for the labels — the admin UI fetches them from
# GET /admin/producers/rejection-presets rather than carrying its own copy,
# so the string the admin clicks, the string persisted to
# producers.rejection_reason, and the string in the producer's email are the
# same object (workflow.md Smell #1: two owners for one fact).
# Insertion order is the display order.
PRODUCER_REJECTION_PRESETS: dict[str, str] = {
    "missing_docs": "מסמכים חסרים / לא קריאים",
    "missing_image": "תמונה ראשית חסרה",
    "incomplete_info": "מידע עסקי לא מלא (כתובת / טלפון / תיאור)",
    "not_eligible": "עסק לא עומד בתנאי הפלטפורמה",
    "other": "אחר (פירוט חופשי)",
}


def _compose_rejection_reason(preset_key: str | None, reason: str) -> str:
    """Join the preset label with the admin's free text into the ONE string
    that is persisted, emailed and sent to the admin WhatsApp line.

    `other` deliberately yields the free text ALONE — its label reads
    "אחר (פירוט חופשי)", which is a UI affordance describing the input, not a
    reason. Prefixing it would put "סיבה: אחר (פירוט חופשי) — ..." in a
    business owner's inbox. Every other preset prefixes its label and appends
    the free text only when the admin actually typed some.

    No preset (legacy callers, and the pre-MEH-226 `{"reason": ...}` body)
    falls through to the free text unchanged.

    PRECONDITION: `preset_key` is None or a key of PRODUCER_REJECTION_PRESETS —
    the route handler 400s on anything else BEFORE calling this. The lookup
    below is therefore deliberately undefended: a KeyError here would mean an
    internal caller skipped that validation, and crashing loudly at the one
    line that noticed beats composing a rejection reason out of a key nobody
    recognises and mailing it to a business owner.
    """
    if preset_key is None or preset_key == "other":
        return reason
    label = PRODUCER_REJECTION_PRESETS[preset_key]
    return f"{label} — {reason}" if reason else label


@router.get("/producers/rejection-presets", response_model=list[RejectionPresetOut])
def list_rejection_presets(user: User = Depends(require_admin)):
    """MEH-226: the reject-modal's radio options. Serving them from the same
    dict the handler composes with is what keeps the admin UI from growing a
    second copy of the Hebrew labels."""
    return [
        {"key": key, "label": label}
        for key, label in PRODUCER_REJECTION_PRESETS.items()
    ]


@router.get("/producers/{producer_id}", response_model=ProducerAdminOut)
@limiter.limit("60/minute")
def admin_get_producer(
    request: Request,
    producer_id: UUID,
    user: User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    """MEH-2072: the single-producer admin read. There was none before, and the
    admin EDIT PAGE was loading the PUBLIC serializer instead.

    ## The bug this closes, measured rather than reasoned

    `app/[locale]/admin/producers/[id]/edit/page.js` fetched `GET /producers/{id}`
    -> `ProducerDetailOut`. That shape carries no admin-only field, by design.
    So `ProducerForm` hydrated every one of them as `""` (its `initial.X ?? ""`
    idiom), and because the form POSTs its whole state back, saving ANY edit
    wrote those blanks over the stored values:

        producer_license_number  '1234567'  ->  ''
        address                  'הרצל 1'   ->  None

    Silent, on every admin save, including a save that only changed the name.
    `producer_license_number` is the regulatory field the whole "licensed
    businesses only" promise rests on, so this was live data loss on the one
    column the promise depends on.

    Two things made it invisible. The form's own comment asserted that the page
    called `GET /admin/producers/{id}` "which exposes the raw
    producer_license_number" — describing this route, which did not exist and
    returned **405**. And the failure has no error surface: the PUT succeeds,
    the page redirects, and the value is simply gone.

    ## Why a new route rather than widening the public one

    Widening `ProducerDetailOut` would publish licence numbers and street
    addresses to every visitor — the exact inversion of the MEH-530 privacy
    precedent. The admin needs a different SHAPE, not more public data.

    ## Route ordering is load-bearing

    Declared AFTER `/producers/pending` (:977) and `/producers/rejection-presets`
    (:1132). FastAPI matches in declaration order, so registering this above
    either of them would swallow both literals into `{producer_id}` and answer
    them with a UUID-parse 422. Do not move it up.

    Mirrors the list endpoint's post-query enrichment exactly — `attach_badge_fields`
    then `_attach_aging` — because a single-row read that skipped them would
    serialise a silent `0` for `business_days_waiting` and default badges, which
    is the "looks measured and is not" failure `_attach_aging`'s own docstring
    warns about.
    """
    producer = (
        db.query(Producer)
        .options(
            joinedload(Producer.categories),
            joinedload(Producer.products),
            joinedload(Producer.delivery_areas),
            # selectinload, matching the list endpoint (:448) — a 4th collection
            # joinedload would widen the existing 3-way cartesian. Needed by
            # attach_badge_fields' pickup_points/offers_pickup derivation.
            selectinload(Producer.locations),
        )
        .filter(Producer.id == producer_id)
        .first()
    )
    if not producer:
        raise HTTPException(status_code=404, detail="בית עסק לא נמצא")
    attach_badge_fields(producer)
    _attach_aging(producer)
    return producer


@router.post("/producers/{producer_id}/reject")
def reject_producer(
    producer_id: UUID,
    body: ProducerRejectIn = Body(default_factory=ProducerRejectIn),
    user: User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    # MEH-226: validate BEFORE mutating — an unknown preset must not leave the
    # producer rejected with an empty reason. Mirrors request_producer_changes'
    # handler-side 400 (admin.py:829).
    preset_key = body.preset_key
    if preset_key is not None and preset_key not in PRODUCER_REJECTION_PRESETS:
        raise HTTPException(status_code=400, detail="סיבת דחייה לא מוכרת")
    reason = (body.reason or "").strip()
    if preset_key == "other" and not reason:
        raise HTTPException(
            status_code=400, detail="יש לפרט את סיבת הדחייה כשנבחר 'אחר'"
        )

    producer = db.query(Producer).filter(Producer.id == producer_id).first()
    if not producer:
        raise HTTPException(status_code=404, detail="בית עסק לא נמצא")
    # MEH-2121: symmetric with approve. Rejecting a draft emails the owner that
    # her application was turned down for an application she never made — the
    # rejection mail and the `rejection_reason` banner both fire on a request
    # that does not exist. No override, same reasoning as approve.
    if producer.status == "draft":
        raise HTTPException(
            status_code=409,
            detail="אי אפשר לדחות עסק שטרם נשלח לבדיקה",
        )
    composed_reason = _compose_rejection_reason(preset_key, reason)
    producer.status = "rejected"
    # MEH-226: the reason is persisted in the SAME commit as the status flip —
    # before this it lived only in the email body, so a rejected owner saw
    # "נדחה" with no reason on her dashboard (the banner reads
    # producer_rejection_reason off GET /auth/me).
    producer.rejection_reason = composed_reason or None
    # MEH-2210: the preset key is the structured reason code. Same dict as the
    # composed text (PRODUCER_REJECTION_PRESETS, validated above), so there is
    # exactly one vocabulary; the owner dashboard branches its copy on it.
    # NULL when a legacy caller sent free text only.
    producer.rejection_reason_code = preset_key
    # MEH-1011: clear any request-changes trail on reject — a rejected producer
    # must not carry a stale "ממתין להשלמה" trail in ProducerAdminOut. Symmetric
    # with approve_producer's clearing.
    producer.requested_changes = None
    producer.changes_requested_at = None
    db.commit()

    # MEH-226: email fires post-commit only — a Resend failure must not roll
    # back a decision the admin already made.
    reason_text = _rejection_reason_suffix(composed_reason)
    producer_user = db.query(User).filter(User.producer_id == producer.id).first()
    if producer_user:
        # MEH-2210: both parts carry the dashboard link — the resubmit CTA
        # lives there. Same link the changes-requested mail builds.
        dashboard_link = f"{settings.frontend_url}/producer/dashboard"
        _send_notification_email(
            producer_user.email,
            f'מהמקור - עדכון לגבי העסק "{producer.name}"',
            _producer_rejected_body(producer.name, composed_reason, dashboard_link),
            html=_producer_rejected_html(
                producer.name, composed_reason, dashboard_link
            ),
        )

    # Notify admin via WhatsApp
    if settings.admin_whatsapp_to:
        _send_whatsapp(
            settings.admin_whatsapp_to,
            f'❌ העסק "{producer.name}" נדחה.{reason_text}',
        )

    return {
        "detail": "Producer rejected",
        "id": str(producer.id),
        "status": producer.status,
        "rejection_reason": producer.rejection_reason,
    }


# MEH-1011: producer request-changes — non-terminal "please complete" path.
# Unlike reject_producer (terminal → status="rejected"), this KEEPS the
# producer pending and records the admin's feedback so the business can fix
# the gap (missing photo / license) and be approved.
@router.post("/producers/{producer_id}/request-changes")
def request_producer_changes(
    producer_id: UUID,
    body: RequestChangesIn,
    user: User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    """Send a completion request to a pending producer. `feedback` is required
    (it becomes the body of the email to the producer) — empty/whitespace → 400,
    mirroring admin_recipes.py:123. Status stays "pending"; no rejection.
    """
    feedback = (body.feedback or "").strip()
    if not feedback:
        raise HTTPException(status_code=400, detail="feedback is required")

    producer = db.query(Producer).filter(Producer.id == producer_id).first()
    if not producer:
        raise HTTPException(status_code=404, detail="בית עסק לא נמצא")

    # MEH-1011: request-changes is a pending-only operation — it deliberately
    # leaves status unchanged, so applying it to an already-decided producer
    # (approved / rejected / inactive) would leave an incoherent record
    # (e.g. approved + non-null requested_changes, which the "ממתין להשלמה"
    # badge keys off). 409 mirrors toggle-status's invalid-transition guard
    # (MEH-769).
    #
    # MEH-2121 correction: this line used to end "reject_producer needs no such
    # guard — it transitions status." That was true before the draft state
    # existed, when every status was a coherent thing to reject. It is false
    # now — reject_producer carries its own `draft` 409 for exactly that reason.
    if producer.status != "pending":
        raise HTTPException(
            status_code=409,
            detail="ניתן לשלוח בקשת השלמה רק לבית עסק בהמתנה לאישור",
        )

    producer.requested_changes = feedback
    # tz-aware (MEH-762 D1, mirrors grant_verified) — the column is TIMESTAMPTZ.
    producer.changes_requested_at = datetime.now(timezone.utc)
    # status intentionally unchanged — stays "pending".
    db.commit()

    p_name = producer.name
    # Env-aware link (mirrors auth_emails.py:159) — staging emails must point
    # at staging, not production. On prod frontend_url == mehamakor.online.
    dashboard_link = f"{settings.frontend_url}/producer/dashboard"

    producer_user = db.query(User).filter(User.producer_id == producer.id).first()
    if producer_user:
        _send_notification_email(
            producer_user.email,
            f'מהמקור — נשאר פרט אחד לפני האישור של "{p_name}"',
            _producer_changes_requested_body(p_name, feedback, dashboard_link),
        )

    # MEH-1051: WhatsApp mirror of the email above (producer_changes_requested_v1,
    # Meta-approved). Post-commit + fail-open — a Meta outage or missing phone
    # must never affect the 200; the function handles skip/raise internally.
    notify_producer_changes_requested(p_name, producer.phone, feedback)

    # Notify admin via WhatsApp
    if settings.admin_whatsapp_to:
        _send_whatsapp(
            settings.admin_whatsapp_to,
            f"📝 נשלחה בקשת השלמה ל-{p_name}",
        )

    return {
        "detail": "בקשת השלמה נשלחה",
        "requested_changes": producer.requested_changes,
        "changes_requested_at": producer.changes_requested_at.isoformat(),
    }


# MEH-762 (ADR-022 public tier contract, Chunk 2): admin stamping for the
# tier-1 "מאומת" badge. The document review itself stays manual off-platform
# (VERIFICATION.md §2/§4) — these endpoints only record the OUTCOME in the DB.
# No auto-stamp on admin-create/import. The legacy is_verified column was
# DROPPED in MEH-766 ch6 (revision d4e7a92c81b5); verified_at is the only
# verification axis. The public verification_tier resolver + exposure landed
# in MEH-762 Chunk 3.
@router.post("/producers/{producer_id}/grant-verified")
def grant_verified(
    producer_id: UUID,
    body: GrantVerifiedIn,
    user: User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    """Stamp the tier-1 verification result after the admin checks the
    qualifying document (license / exemption / cosmetics — VERIFICATION.md
    §2). Re-grant overwrites verified_at + verification_doc_type (the legit
    correction path alongside revoke-verified).
    # REUSES: admin_kashrut.py:75 — admin stamps a verification timestamp.
    """
    producer = db.query(Producer).filter(Producer.id == producer_id).first()
    if not producer:
        raise HTTPException(status_code=404, detail="בית עסק לא נמצא")
    # tz-aware (MEH-762 D1, mirrors MEH-759) — NOT naive utcnow; the column
    # is TIMESTAMPTZ. Public exposure is date-granularity only (Chunk 3).
    producer.verified_at = datetime.now(timezone.utc)
    producer.verification_doc_type = body.doc_type
    db.commit()
    return {
        "detail": "תג מאומת הוענק",
        "verified_at": producer.verified_at.isoformat(),
        "verification_doc_type": producer.verification_doc_type,
    }


@router.post("/producers/{producer_id}/revoke-verified")
def revoke_verified(
    producer_id: UUID,
    user: User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    """Clear the tier-1 "מאומת" stamp (mistake correction). Idempotent —
    clearing an already-unverified producer is a no-op success. (The legacy
    is_verified column was dropped in MEH-766 ch6.)
    """
    producer = db.query(Producer).filter(Producer.id == producer_id).first()
    if not producer:
        raise HTTPException(status_code=404, detail="בית עסק לא נמצא")
    producer.verified_at = None
    producer.verification_doc_type = None
    db.commit()
    return {"detail": "תג מאומת הוסר"}


# MEH-1406: admin home-products moderation endpoints removed from the live
# surface (was: GET /home-products/hidden, POST .../restore, DELETE .../{id},
# GET /home-products/flagged, POST .../approve, POST .../remove). The
# consumer home-cook feature was retired per brand LOCK (licensed
# businesses only), so its admin moderation queue is unmounted too. The
# HomeProduct model/schemas/tables are retained (no Alembic) — the /admin/stats
# counts below still read them; only the writable/queue endpoints are gone.


# MEH-587: admin Recipe endpoints removed (chunk 0/4) — see
# backend/alembic/versions/20260515_1430_d7e3c9a82f5b_meh_587_remove_zombie_recipes.py.


# --- MEH-53: Instagram story card ---


@router.post("/producers/{producer_id}/story-card")
def upload_story_card(
    producer_id: UUID,
    body: StoryCardUploadRequest,
    user: User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    """Accept base64 JPEG canvas export, upload to Cloudinary, persist URL."""
    producer = db.query(Producer).filter(Producer.id == producer_id).first()
    if not producer:
        raise HTTPException(status_code=404, detail="בית עסק לא נמצא")

    # Strip the data-URI prefix to get raw base64
    data_uri = body.image_data
    if "," in data_uri:
        data_uri = data_uri.split(",", 1)[1]

    import base64

    try:
        raw = base64.b64decode(data_uri)
    except Exception:
        raise HTTPException(status_code=400, detail="נתוני תמונה לא תקינים")

    if not settings.cloudinary_cloud_name:
        # Dev fallback
        return {"url": f"/placeholder-image.png?id={producer_id}", "saved": False}

    try:
        import cloudinary
        import cloudinary.uploader

        cloudinary.config(
            cloud_name=settings.cloudinary_cloud_name,
            api_key=settings.cloudinary_api_key,
            api_secret=settings.cloudinary_api_secret,
        )
        result = cloudinary.uploader.upload(
            raw,
            folder=f"mehamakor/producers/{producer_id}",
            public_id="story-card",
            resource_type="image",
            overwrite=True,
            format="jpg",
        )
        url = result["secure_url"]
        producer.story_card_url = url
        db.commit()
        return {"url": url, "saved": True}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Upload failed: {str(e)}")


# --- Stats ---
@router.get("/stats")
def get_stats(user: User = Depends(require_admin), db: Session = Depends(get_db)):
    return {
        "total_producers": db.query(Producer).count(),
        "pending_producers": db.query(Producer)
        .filter(Producer.status == "pending")
        .count(),
        "approved_producers": db.query(Producer)
        .filter(Producer.status == "approved")
        .count(),
        "total_users": db.query(User).count(),
        "total_home_products": db.query(HomeProduct)
        .filter(HomeProduct.is_active.is_(True))
        .count(),
        "hidden_home_products": db.query(HomeProduct)
        .filter(HomeProduct.is_hidden.is_(True))
        .count(),
    }


_DATA_GOV_URL = (
    "https://data.gov.il/api/3/action/datastore_search"
    "?resource_id=d4901968-dad3-4845-a9b0-a57d027f11ab&limit=1500"
)


# MEH-2241: this handler used to carry an inline copy of the parser in
# `scripts/seed_cities.py` — same guessed column names, same silent outcome.
# On Railway staging that script reported `Received 1272 records`, inserted
# nothing and exited 0, because the resource publishes `שם_ישוב` (one yod)
# while both parsers read `שם_יישוב` (two yods). PR #3288 fixed the script by
# discovering the column from the response; this route now calls the same
# function rather than holding a second copy of the bug. Two owners for one
# parse is the drift `.claude/rules/workflow.md` Smell #1 names, and here it
# was not hypothetical — fixing one would have left the other broken.
#
# The rationale lives in a comment, not the docstring: FastAPI publishes a
# route's docstring as the endpoint `description` in `backend/openapi.json`,
# which is a committed artifact that `frontend/lib/generated/` is derived
# from. Internal history does not belong in the API contract, and putting it
# there drags two generated files into every such diff.
@router.post("/seed-cities", status_code=200)
def seed_cities(
    user: User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    """Idempotent: fetch Israeli localities from data.gov.il and upsert into cities table."""
    # Imported here, not at module scope: `scripts.seed_cities` calls
    # `logging.basicConfig` when imported (it is written to be run as a
    # script), and the API process must not inherit that at boot merely
    # because this route exists.
    from scripts.seed_cities import LocalityParseError, parse_localities

    try:
        resp = httpx.get(_DATA_GOV_URL, timeout=30)
        resp.raise_for_status()
    except httpx.HTTPError as exc:
        raise HTTPException(status_code=502, detail=f"data.gov.il fetch failed: {exc}")

    try:
        cities = parse_localities(resp.json())
    except LocalityParseError as exc:
        # The message names the keys the records DID carry — the one piece of
        # evidence nobody could get from the sandbox. Surfacing it beats a 200
        # that says "seeded 0" and explains nothing.
        raise HTTPException(
            status_code=502,
            detail=f"הרשומות התקבלו אך לא נמצאה בהן עמודת שם יישוב: {exc}",
        )

    inserted = 0
    for city in cities:
        result = db.execute(
            text(
                "INSERT INTO cities (name_he, lat, lng) VALUES (:name_he, :lat, :lng)"
                " ON CONFLICT (name_he) DO NOTHING"
            ),
            city,
        )
        inserted += result.rowcount
    db.commit()
    return {"seeded": inserted}


# MEH-2027: the three PRODUCER-facing bodies below were inline f-strings inside
# their route handlers, which put them out of reach of the copy contract in
# tests/test_meh1965_email_copy_contract.py — rendering one meant standing up a
# request, a DB row and an authenticated admin. They are pure functions of their
# arguments so the contract can render them directly. Behaviour is unchanged:
# the strings are byte-identical to the f-strings they replace.
#
# These go to producer_user.email — the business owner, NOT an admin — so they
# are brand touchpoints and the contract's four axes (absolute links, no
# masculine address to the reader, RTL on any HTML part, real text fallback)
# apply to them in full.


def _rejection_reason_suffix(reason: str) -> str:
    """The optional reason tail, shared by the email body and the admin
    WhatsApp line so the two cannot drift apart (workflow.md Smell #1).

    MEH-226: the label was "סיבת הדחייה" and is now "הסיבה", per the approved
    email copy. The admin WhatsApp line changes with it BY DESIGN — one owner
    is the whole point of this helper, and splitting it to keep the WhatsApp
    wording frozen would rebuild exactly the drift it exists to prevent.

    Still returns "" for an empty reason, so the rejected-without-a-reason
    email omits the line rather than shipping a dangling "הסיבה:".
    """
    return f"\nהסיבה: {reason}" if reason else ""


# MEH-2134: LOCKED copy, approved by Sapir 20/08/2026, shipped verbatim.
_APPROVED_COMMUNITY_BLOCK = (
    "פתחנו קבוצת עדכונים בוואטסאפ לבתי עסק שאושרו במהמקור — שווקים ואירועים "
    "לפני כולם, ומה חדש באתר. פעם-פעמיים בחודש, רק אנחנו כותבות שם, ואפשר "
    "לצאת בכל רגע:"
)

# MEH-2151: LOCKED copy, approved by Sapir 21/08/2026, shipped verbatim.
# Labels only — the URL is appended on its own line beneath each, mirroring
# the community block's `label:\n{url}` shape so all three read alike.
_APPROVED_VIEW_PAGE_LABEL = "ככה העמוד שלך נראה ללקוחות:"
_APPROVED_DASHBOARD_LABEL = "לעדכון פרטים, תמונות ומוצרים — לוח הבקרה:"
_APPROVED_HTML_BUTTON_LABEL = "לצפייה בעמוד העסק"
_APPROVED_HTML_DASHBOARD_LABEL = "לוח הבקרה — עדכון פרטים ומוצרים"


def _approval_links(slug: str | None) -> tuple[str, str]:
    """The two absolute URLs the approval email carries: (page, dashboard).

    Shared by the text and HTML builders deliberately. MEH-2151 requires BOTH
    parts to carry BOTH links, so composing the URLs twice would create two
    owners for one fact — the drift `workflow.md` Smell #1 names — and the
    failure would be silent: an email whose button and whose text line point at
    different places still renders, still sends, and still looks correct in a
    diff. This is not a shared layout module (the over-engineering guard's
    target); it returns two strings and knows nothing about either part's
    markup.

    The page URL is "" for a falsy or whitespace-only slug, which is what lets
    each caller drop its view-page block whole. `.strip()` before the guard,
    not after — same reasoning as the community invite below: a whitespace-only
    value is truthy, so an unstripped read would emit a label above a URL of
    "/p/   ". Mirrors the falsy-guard pattern already used for
    `whatsapp_community_invite_url`.

    Env-aware host (mirrors the changes-requested body's `dashboard_link`):
    staging emails must point at staging, not production.
    """
    clean_slug = (slug or "").strip()
    page_url = f"{settings.frontend_url}/p/{clean_slug}" if clean_slug else ""
    dashboard_url = f"{settings.frontend_url}/producer/dashboard"
    return page_url, dashboard_url


def _producer_approved_body(name: str, slug: str | None) -> str:
    """MEH-2134: copy approved by Sapir 20/08/2026, shipped verbatim.

    Signed `ספיר שנפ | מייסדת` in the first person, matching
    `pending_nudge.py:149` and all four `onboarding_followup.py` steps —
    one mailbox, one sender. `_producer_rejected_body` deliberately keeps the
    institutional sign-off: an institutional "no" is kinder than a personal
    one, and that asymmetry is a decision, not an oversight. (The literal is
    spelled out nowhere in this docstring so that the sign-off census below
    counts signatures, not prose.)

    Census of the institutional sign-off in this file: **2** — the rejection
    body and the changes-requested body, both untouched here. MEH-2134's DoD
    predicted 2 → 1, which is wrong twice over: the pre-change count was 3
    (this body plus those two, verified on `origin/staging`), and 1 is
    unreachable because the same ticket requires `_producer_changes_requested_body`
    byte-identical. This function is the only one MEH-2134 moves off it.

    The community invite rides in THIS email rather than the registration
    form (just-in-time beats upfront, and the group is for *approved*
    businesses) and rather than the `producer_approved_v1` WhatsApp template
    (Meta approved it with one parameter and no buttons; adding a URL button
    means re-approval — MEH-509 PR1 got a 400 for exactly that).

    An unset `whatsapp_community_invite_url` drops the paragraph AND the URL
    line together, with no dangling label and no blank-line artifact. That is
    what lets this merge before the link exists in Railway (Phase C).
    Falsy-guard pattern mirrors `settings.admin_whatsapp_to` at :971.

    MEH-2151 added the two link blocks and the `slug` parameter. Order is
    greeting → approval → view-page → dashboard → community → signature: the
    community block moved BELOW the links because it is the secondary action,
    and its text is byte-identical (MEH-2134 LOCKED — the census above is why
    that matters). The mail previously announced "הפרופיל שלך כעת גלוי" and
    then offered no way to look at it; the view-page link is that proof.

    THREE optional blocks now share one shape, and the degradation is the
    reason to keep it that way: with no slug AND no invite the body is
    greeting → approval → dashboard → signature, still with no dangling label.
    Only the dashboard block is unconditional — its URL needs no input beyond
    `settings.frontend_url`, so there is no state in which it cannot be built.
    """
    # `.strip()` before the guard, not after: a Railway value of "  " is
    # truthy, so the raw read would emit the paragraph above an empty line —
    # the dangling-label failure this function exists to avoid, reached by the
    # one input nobody types deliberately.
    invite_url = settings.whatsapp_community_invite_url.strip()
    community = f"\n{_APPROVED_COMMUNITY_BLOCK}\n{invite_url}\n" if invite_url else ""
    page_url, dashboard_url = _approval_links(slug)
    # Same `\n{label}\n{url}\n` shape as `community`, so each block that is
    # present contributes exactly one blank line before it and the signature's
    # own leading "\n" closes the last one — which is why an omitted block
    # leaves no double blank line and no dangling label.
    view_page = f"\n{_APPROVED_VIEW_PAGE_LABEL}\n{page_url}\n" if page_url else ""
    dashboard = f"\n{_APPROVED_DASHBOARD_LABEL}\n{dashboard_url}\n"
    return (
        f'היי,\n\nהעסק שלך "{name}" אושר במהמקור! '
        f"הפרופיל שלך כעת גלוי לכל המשתמשים באתר.\n"
        f"{view_page}"
        f"{dashboard}"
        f"{community}"
        f"\nספיר שנפ\nמייסדת | מהמקור\n{SITE_DOMAIN}"
    )


def _producer_approved_html(name: str, slug: str | None) -> str:
    """MEH-2151: the HTML twin of `_producer_approved_body` — same copy, same order.

    WHY AN HTML PART EXISTS AT ALL, since the text body was already readable:
    Gmail does not infer direction from content, so a plain-text Hebrew line
    ending in a period renders that period at the START of the line — observed
    21/08 on a real approval mail (".גלוי"). `dir="rtl"` on the document plus
    `direction:rtl` on the containing elements is the fix, and the copy contract
    (`tests/test_meh1965_email_copy_contract.py`) asserts BOTH: the attribute
    alone still leaves an inline BiDi run (a Latin business name mid-sentence)
    to the client's guess.

    Structure and palette follow `routers/marketing.py:_send_newsletter_welcome`
    — table layout, inline CSS only, no <style> block and no external sheet,
    because Gmail strips both. No emoji, unlike that precedent: the text twin
    carries none and the two parts must read alike.

    THE CARD IS `width="100%"` CAPPED BY `max-width`, NEVER `width="560"`. The
    HTML width ATTRIBUTE beats the CSS `max-width` beside it, so the two
    together are not belt-and-braces — the attribute simply wins and the card
    is 560px wide on a 375px screen. Measured with Playwright on staging
    `88e9717f`: `scrollWidth=560` against `innerWidth=375`, the button label
    clipped to «עסק». The second-order cost is the one that matters: a client
    that shrink-to-fits instead of scrolling (Gmail's common behaviour) scales
    560→375, and the 46px button lands at ~31px physical — under the 44px
    minimum tap target it passes at CSS scale. `margin:0 auto` keeps it centred
    once the width is fluid.

    KNOWN LIMIT, stated rather than papered over: Outlook's Word engine ignores
    `max-width`, so there the card renders full-width instead of capped. The
    bulletproof fix is an `<!--[if mso]>` fixed-width wrapper, which this
    ticket did not authorise and which the newsletter precedent does not carry
    either. Gmail and Apple Mail — the ticket's smoke targets — honour it.

    The primary button centres with `align="center"` on its own table: an HTML
    attribute, chosen over `margin:auto` because Outlook ignores auto margins
    on a table box. It sits inside a `text-align:right` cell, which does not
    move a block-level table on its own.

    EVERY interpolated value goes through `escape()` — all 11 sites, including
    the LOCKED copy constants and `SITE_DOMAIN`, none of which carries an HTML
    metacharacter today (measured). That is exactly why escaping them is not
    optional: a guarantee that holds only because of a property of the current
    value is not a guarantee, and the next approved copy edit is where it stops
    holding.

    That sentence was FALSE when first written — it claimed "every" while three
    constants were interpolated raw, and the CI reviewer caught it on PR #3054.
    A docstring asserting a property is the artifact least likely to be
    re-checked (`.claude/rules/testing.md`), so the claim is no longer left on
    its own: `test_meh2151_approval_email_cta.py` parses this function's source
    and fails on any interpolation site that is not wrapped, which is what makes
    the sentence above checkable instead of merely stated. `name` is
    owner-supplied free text that reaches this function unfiltered, so a
    business name containing `<` or `"` would break out of the markup it lands
    in — harmless in the text part, an injection in this one. `slug` is
    system-minted (MEH-1817) and already URL-safe, but it is escaped on the same
    line rather than trusted, so the guarantee lives here instead of depending
    on a property enforced two modules away.

    A falsy slug drops the primary button whole, exactly as the text twin drops
    its view-page block: the mail then leads with the dashboard link rather than
    showing a button that goes nowhere.
    """
    page_url, dashboard_url = _approval_links(slug)
    safe_name = _html_escape(name)
    invite_url = settings.whatsapp_community_invite_url.strip()

    button = (
        f'<table align="center" cellpadding="0" cellspacing="0" '
        f'style="margin:0 auto 24px;">'
        f'<tr><td style="background:#2e6853;border-radius:8px;">'
        f'<a href="{_html_escape(page_url)}" style="display:inline-block;'
        f"padding:14px 32px;color:#ffffff;font-size:16px;font-weight:bold;"
        f'text-decoration:none;direction:rtl;">'
        f"{_html_escape(_APPROVED_HTML_BUTTON_LABEL)}</a></td></tr></table>"
        if page_url
        else ""
    )
    community = (
        f'<p style="color:#3a3a3a;font-size:15px;line-height:1.8;'
        f'margin:0 0 24px;direction:rtl;">'
        f"{_html_escape(_APPROVED_COMMUNITY_BLOCK)}<br>"
        f'<a href="{_html_escape(invite_url)}" '
        f'style="color:#2e6853;word-break:break-all;">'
        f"{_html_escape(invite_url)}</a></p>"
        if invite_url
        else ""
    )

    return (
        "<!DOCTYPE html>\n"
        '<html dir="rtl" lang="he">\n'
        '<head><meta charset="UTF-8">'
        '<meta name="viewport" content="width=device-width,initial-scale=1">'
        "</head>\n"
        '<body style="margin:0;padding:0;background:#F5F0E8;'
        'font-family:Arial,Helvetica,sans-serif;direction:rtl;">\n'
        '<table width="100%" cellpadding="0" cellspacing="0" '
        'style="background:#F5F0E8;padding:32px 0;"><tr><td align="center">\n'
        '<table width="100%" cellpadding="40" cellspacing="0" '
        'style="background:#ffffff;border-radius:12px;text-align:right;'
        'direction:rtl;max-width:560px;margin:0 auto;"><tr>\n'
        '<td style="text-align:right;direction:rtl;">\n'
        '<p style="color:#3a3a3a;font-size:15px;line-height:1.8;'
        'margin:0 0 16px;direction:rtl;">היי,</p>\n'
        '<p style="color:#1C1A17;font-size:17px;line-height:1.8;'
        'margin:0 0 24px;direction:rtl;">'
        f'העסק שלך "{safe_name}" אושר במהמקור! '
        "הפרופיל שלך כעת גלוי לכל המשתמשים באתר.</p>\n"
        f"{button}\n"
        '<p style="margin:0 0 24px;direction:rtl;">'
        f'<a href="{_html_escape(dashboard_url)}" style="color:#2e6853;'
        'font-size:15px;text-decoration:underline;direction:rtl;">'
        f"{_html_escape(_APPROVED_HTML_DASHBOARD_LABEL)}</a></p>\n"
        f"{community}\n"
        '<hr style="border:none;border-top:1px solid #e5e0d8;margin:0 0 20px;">\n'
        '<p style="color:#3a3a3a;font-size:15px;line-height:1.8;'
        'margin:0;direction:rtl;">'
        f"ספיר שנפ<br>מייסדת | מהמקור<br>{_html_escape(SITE_DOMAIN)}</p>\n"
        "</td></tr></table>\n"
        "</td></tr></table>\n"
        "</body>\n"
        "</html>"
    )


def _rejected_dashboard_link(dashboard_link: str | None) -> str:
    """MEH-2210: the owner dashboard is where the resubmit CTA lives. Computed
    lazily so the MEH-1965 copy corpus can keep calling the body builders with
    two arguments; the handler passes the explicit value it already builds for
    the changes-requested mail (`admin.py` request_producer_changes)."""
    return dashboard_link or f"{settings.frontend_url}/producer/dashboard"


def _single_line(value: str) -> str:
    """Collapse any newline in an owner-supplied value to a single space.

    `sanitize_text` (services/sanitization.py:14) runs bleach with `tags=[]`
    and then `.strip()` — bleach removes markup, `.strip()` trims the ends,
    and neither touches an interior `\n`. So a business name really can carry
    line breaks into a plain-text mail body, where the HTML twin's
    `_html_escape` gives no protection because the text part is not markup.

    In the rejection mail the greeting is the first line and the signature the
    last, so an embedded newline lets the name open lines of its own beneath
    the greeting — a self-addressed mail, hence Minor and not a security fix,
    but the body should render as the copy Sapir approved regardless.

    Reported by the reviewer on the MEH-2210 chunk-C PR against this function.
    The same interpolation exists in `_producer_approved_body` (:1777) and
    `_producer_changes_requested_body` (:2015); both are OUTSIDE this PR's
    diff and are deliberately left alone here rather than widening it — this
    helper is what makes each of them a one-line change.
    """
    return " ".join(value.split())


def _producer_rejected_body(
    name: str, reason: str, dashboard_link: str | None = None
) -> str:
    """MEH-226: copy approved by Sapir 14.08.2026 — greeting, decision and the
    "הסיבה:" tail are shipped verbatim.

    MEH-2210 replaced the recovery line. The MEH-226 sentence ("אפשר לתקן את
    הפרטים בלוח הבקרה ולהשיב למייל הזה — ונבחן את הבקשה מחדש") was the
    conditional half of that ruling: it said "reply to this email" BECAUSE the
    resubmit flow did not exist — `request_producer_review` answered a rejected
    owner with 409. Chunk A of MEH-2210 opened that door (rejected → pending,
    three times), so the line now points at the flow that exists, with the
    card's own copy and the dashboard link the changes-requested mail already
    carries. `tests/test_meh226_rejection_reason.py` holds both halves: the
    owner can still edit, and the retired "הגישי שוב מהדף האישי" stays out.

    The reason tail is the COMPOSED text (`_compose_rejection_reason`) — the
    preset label plus the admin's free text — so the "reason line by code" the
    card asks for is already in it; a second line keyed on the code would print
    the same label twice.
    """
    link = _rejected_dashboard_link(dashboard_link)
    return (
        f"שלום {_single_line(name)},\n\n"
        f"תודה על הבקשה להצטרף למהמקור. בשלב זה לא אישרנו אותה."
        f"{_rejection_reason_suffix(reason)}\n\n"
        f"אפשר לתקן ולשלוח שוב מלוח הבקרה: {link}\n\n"
        f"בברכה,\nצוות מהמקור"
    )


def _producer_rejected_html(
    name: str, reason: str, dashboard_link: str | None = None
) -> str:
    """MEH-2210: the HTML twin of `_producer_rejected_body` — same copy, same
    order, same reason. Mirrors `_producer_approved_html` (MEH-2151): Gmail
    does not infer direction from content, so the plain-text Hebrew body
    renders its trailing period at the START of the line; `dir="rtl"` on the
    document plus `direction:rtl` on every containing element is the fix.
    Every producer-controlled value is escaped; the reason is the admin's
    composed text and is escaped too — it is rendered, never trusted.
    """
    # MEH-2210: _single_line first, so the two twins agree on what an
    # owner-supplied name may do. Reviewer's finding, and their own
    # caveat is right — a newline is invisible here because HTML
    # collapses whitespace, so this buys symmetry rather than a fix.
    # That is worth one call: the text twin's guarantee is only
    # readable as a rule if both paths carry it.
    safe_name = _html_escape(_single_line(name))
    link = _rejected_dashboard_link(dashboard_link)
    reason_block = (
        (
            '<p style="color:#3a3a3a;font-size:15px;line-height:1.8;'
            'margin:0 0 24px;direction:rtl;">'
            f"הסיבה: {_html_escape(reason)}</p>\n"
        )
        if reason
        else ""
    )
    return (
        "<!DOCTYPE html>\n"
        '<html dir="rtl" lang="he">\n'
        '<head><meta charset="UTF-8">'
        '<meta name="viewport" content="width=device-width,initial-scale=1">'
        "</head>\n"
        '<body style="margin:0;padding:0;background:#F5F0E8;'
        'font-family:Arial,Helvetica,sans-serif;direction:rtl;">\n'
        '<table width="100%" cellpadding="0" cellspacing="0" '
        'style="background:#F5F0E8;padding:32px 0;"><tr><td align="center">\n'
        '<table width="100%" cellpadding="40" cellspacing="0" '
        'style="background:#ffffff;border-radius:12px;text-align:right;'
        'direction:rtl;max-width:560px;margin:0 auto;"><tr>\n'
        '<td style="text-align:right;direction:rtl;">\n'
        '<p style="color:#3a3a3a;font-size:15px;line-height:1.8;'
        f'margin:0 0 16px;direction:rtl;">שלום {safe_name},</p>\n'
        '<p style="color:#1C1A17;font-size:17px;line-height:1.8;'
        'margin:0 0 24px;direction:rtl;">'
        "תודה על הבקשה להצטרף למהמקור. בשלב זה לא אישרנו אותה.</p>\n"
        f"{reason_block}"
        '<p style="margin:0 0 24px;direction:rtl;">'
        "אפשר לתקן ולשלוח שוב מלוח הבקרה: "
        f'<a href="{_html_escape(link)}" style="color:#2e6853;'
        'font-size:15px;text-decoration:underline;direction:rtl;">'
        f"{_html_escape(link)}</a></p>\n"
        '<hr style="border:none;border-top:1px solid #e5e0d8;margin:0 0 20px;">\n'
        '<p style="color:#3a3a3a;font-size:15px;line-height:1.8;'
        'margin:0;direction:rtl;">'
        f"בברכה,<br>צוות מהמקור<br>{_html_escape(SITE_DOMAIN)}</p>\n"
        "</td></tr></table>\n"
        "</td></tr></table>\n"
        "</body>\n"
        "</html>"
    )


def _producer_changes_requested_body(
    name: str, feedback: str, dashboard_link: str
) -> str:
    return (
        f"שלום,\n\n"
        f'הבקשה לרישום "{name}" במהמקור נבדקה. כדי שנוכל לאשר ולפרסם את '
        f"בית העסק, נשאר להשלים:\n\n"
        f"{feedback}\n\n"
        f"אפשר להשלים את הפרטים בלוח הבקרה: {dashboard_link}\n"
        f"לאחר ההשלמה נמשיך בתהליך האישור ונעדכן אתכם.\n\n"
        f"בברכה,\nצוות מהמקור"
    )


def _send_notification_email(
    to_email: str, subject: str, body: str, html: str | None = None
):
    """MEH-2151: `html` is OPTIONAL and defaults to None.

    That default is what keeps the three pre-existing callers (rejected,
    changes-requested, and the approval site before this ticket) byte-identical
    on the wire: `send_email` only adds `params["html"]` when the argument is
    truthy (`services/email.py`), so an omitted argument produces the exact
    payload it produced before this parameter existed. Same shape, and same
    reasoning, as `reply_to` two parameters down in `send_email` (MEH-2112).
    """
    send_email(to_email, subject, body, html=html)


def _send_whatsapp(to: str, body: str):
    """Send WhatsApp admin notification via Meta Cloud API.

    MEH-508: send_text is fail-open (False on missing config or HTTP error,
    no exception raised), so the previous try/except + configured-check
    collapse to a single call. Service-level logger emits the warning.
    """
    send_text(to, body)
