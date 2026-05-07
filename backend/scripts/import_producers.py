"""Standalone CLI script to import producers from an Excel file.

Usage:
    python -m scripts.import_producers <path-to-xlsx> [--dry-run]

Excel column mapping is documented in app/services/producer_import.py.
"""

import argparse
import sys
from pathlib import Path

# Allow running as `python scripts/import_producers.py` from backend/
ROOT = Path(__file__).resolve().parent.parent
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from openpyxl import load_workbook  # noqa: E402

from app.database import SessionLocal  # noqa: E402
from app.services.producer_import import import_rows  # noqa: E402


def main():
    parser = argparse.ArgumentParser(description="Import producers from Excel.")
    parser.add_argument("file", help="Path to .xlsx file")
    parser.add_argument("--dry-run", action="store_true", help="Preview without saving")
    args = parser.parse_args()

    file_path = Path(args.file)
    if not file_path.exists():
        sys.exit(f"File not found: {file_path}")

    wb = load_workbook(file_path, data_only=True)
    ws = wb.active
    rows = [list(r) for r in ws.iter_rows(min_row=2, values_only=True)]

    db = SessionLocal()
    try:
        result = import_rows(db, rows, dry_run=args.dry_run)
    finally:
        db.close()

    print(f"\n=== Import {'PREVIEW' if args.dry_run else 'COMPLETE'} ===")
    print(f"  Imported: {result['imported']}")
    print(f"  Skipped:  {result['skipped']}")
    print(f"  Errors:   {result['errors']}")
    print()

    error_rows = [r for r in result["rows"] if r["errors"]]
    if error_rows:
        print("--- Errors ---")
        for r in error_rows:
            print(
                f"  Row {r['row_number']} ({r['data'].get('name', '?')}): {', '.join(r['errors'])}"
            )

    warning_rows = [r for r in result["rows"] if r["warnings"] and not r["errors"]]
    if warning_rows:
        print("\n--- Warnings ---")
        for r in warning_rows:
            print(
                f"  Row {r['row_number']} ({r['data'].get('name', '?')}): {', '.join(r['warnings'])}"
            )


if __name__ == "__main__":
    main()
